import re
from datetime import datetime
import logging
import openpyxl
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

API_URL = "http://localhost:8000/api/employees"

VALIDATORS = {
    "employee_id": None,  # assigned after function definitions
    "name":        None,
    "email":       None,
    "department":  None,
    "salary":      None,
    "join_date":   None,
}


def validate_employee_id(value, seen_ids):
    if value is None:
        return "Employee ID is required"
    if value in seen_ids:
        return "Employee ID must be unique"
    seen_ids.add(value)
    return None


def validate_name(value, **_):
    if not value or not (1 <= len(str(value)) <= 50):
        return "name must be 1-50 characters"
    return None


def validate_email(value, **_):
    if not value or not re.match(r"[^@]+@[^@]+\.[^@]+", str(value)):
        return "invalid email address"
    return None


def validate_department(value, **_):
    if not value or str(value).strip() == "":
        return "department is required"
    return None


def validate_salary(value, **_):
    if value is None or not isinstance(value, int) or isinstance(value, bool) or value <= 0:
        return "salary must be a positive integer"
    return None


def validate_join_date(value, **_):
    if value is None:
        return "join_date is required"
    try:
        datetime.strptime(str(value), "%Y-%m-%d")
    except ValueError:
        return "invalid date format. Expected YYYY-MM-DD"
    return None


VALIDATORS = {
    "employee_id": validate_employee_id,
    "name":        validate_name,
    "email":       validate_email,
    "department":  validate_department,
    "salary":      validate_salary,
    "join_date":   validate_join_date,
}


def submit_row(record: dict, row_idx: int) -> bool:
    try:
        response = requests.post(API_URL, json=record, timeout=15)
        if response.status_code == 201:
            logging.info(f"Row {row_idx}: Successfully created employee {record['employee_id']}")
            return True
        else:
            logging.error(f"Row {row_idx}: Failed to create employee {record['employee_id']}: {response.status_code} {response.text}")
            return False
    except requests.RequestException as e:
        logging.error(f"Row {row_idx}: Failed to create employee {record['employee_id']}: {str(e)}")
        return False


def main():
    wb = openpyxl.load_workbook('interview_employee_data.xlsx')
    ws = wb["Employee Data"]

    headers = [cell.value for cell in ws[1]]
    seen_ids = set()
    failed_rows = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(headers, row))
        errors = []
        for field, validator in VALIDATORS.items():
            error = validator(record.get(field), seen_ids=seen_ids)
            if error:
                errors.append(f"{field}: {error}")
                print(f"Row {idx}: {field} - {error}")
        if errors:
            logging.warning(f"Row {idx} validation failed - {'; '.join(errors)}")
            continue

        success = submit_row(record, idx)
        if not success:
            failed_rows.append(idx)

    if failed_rows:
        logging.error(f"Failed rows: {failed_rows}")


if __name__ == "__main__":
    main()
